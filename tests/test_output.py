"""Tests for output formatting."""
import csv
import io
from pathlib import Path

from rich.console import Console

from competitor_analysis.models import CompetitorRow
from competitor_analysis.output.table import render_table, _followers_cell, _structure_cell
from competitor_analysis.output.export import export_csv, _flatten


def test_followers_cell_shows_non_na(sample_row: CompetitorRow):
    cell = _followers_cell(sample_row)
    assert "Instagram: 5.2K" in cell
    assert "Linkedin: 800" in cell
    # N/A platforms should be omitted
    assert "Facebook" not in cell


def test_structure_cell(sample_row: CompetitorRow):
    cell = _structure_cell(sample_row)
    assert "Website" in cell
    assert "Landing page" in cell
    assert "E-book" not in cell


def test_render_table_no_crash(sample_row: CompetitorRow):
    console = Console(file=io.StringIO(), width=200)
    render_table([sample_row], console=console)
    output = console.file.getvalue()
    assert "Test Agency" in output
    assert "Competitor Analysis" in output


def test_flatten_has_all_keys(sample_row: CompetitorRow):
    flat = _flatten(sample_row)
    assert "name" in flat
    assert "followers_instagram" in flat
    assert "followers_tiktok" in flat
    assert "has_website" in flat
    assert "why_competitor" in flat


def test_export_csv_roundtrip(sample_row: CompetitorRow, tmp_path: Path):
    out = tmp_path / "test.csv"
    export_csv([sample_row], out)
    with open(out, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    assert len(rows) == 1
    assert rows[0]["name"] == "Test Agency"
    assert rows[0]["followers_instagram"] == "5.2K"


def test_export_excel_roundtrip(sample_row: CompetitorRow, tmp_path: Path):
    import openpyxl
    out = tmp_path / "test.xlsx"
    export_excel_fn = __import__(
        "competitor_analysis.output.export", fromlist=["export_excel"]
    ).export_excel
    export_excel_fn([sample_row], out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    # Row 1 is header, row 2 is first data row
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "Name" in headers
    name_col = headers.index("Name") + 1
    assert ws.cell(2, name_col).value == "Test Agency"
