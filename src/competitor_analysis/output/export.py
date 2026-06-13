from __future__ import annotations

import csv
from pathlib import Path

from competitor_analysis.models import CompetitorRow, MarketingStrategy

_PLATFORMS = ["instagram", "facebook", "linkedin", "youtube", "tiktok"]


def _flatten(row: CompetitorRow) -> dict[str, str]:
    followers = {
        f"followers_{p}": row.kpis.follower_count.get(p, "N/A")
        for p in _PLATFORMS
    }
    structure = row.kpis.structure
    return {
        "name": row.name,
        "description": row.description,
        "activities": row.activities,
        "active_since": row.active_since,
        "social_instagram": row.social_profiles.get("instagram", "N/A"),
        "social_facebook": row.social_profiles.get("facebook", "N/A"),
        "social_linkedin": row.social_profiles.get("linkedin", "N/A"),
        "social_youtube": row.social_profiles.get("youtube", "N/A"),
        "social_tiktok": row.social_profiles.get("tiktok", "N/A"),
        "website_and_links": " | ".join(row.website_and_links),
        **followers,
        "engagement": row.kpis.interaction_score,
        "has_website": str(structure.get("website", False)),
        "has_landing_page": str(structure.get("landing_page", False)),
        "has_ebook": str(structure.get("ebook", False)),
        "has_freebie": str(structure.get("freebie", False)),
        "multi_platform": str(structure.get("multi_platform", False)),
        "why_competitor": row.why_competitor,
    }


def export_csv(rows: list[CompetitorRow], output_path: Path) -> None:
    """Export competitor rows to a CSV file."""
    if not rows:
        return
    flat_rows = [_flatten(r) for r in rows]
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(flat_rows[0].keys()))
        writer.writeheader()
        writer.writerows(flat_rows)


def strategy_to_markdown(strategy: MarketingStrategy) -> str:
    """Render a MarketingStrategy as a clean Italian Markdown document."""
    lines: list[str] = [
        "# Strategia di Marketing",
        "",
        f"**Obiettivo:** {strategy.objective}",
        "",
    ]

    if strategy.summary:
        lines += ["## Sintesi strategica", "", strategy.summary, ""]

    if strategy.positioning:
        lines += ["## Posizionamento", "", strategy.positioning, ""]

    if strategy.differentiation:
        lines += ["## Differenziazione vs competitor", ""]
        for item in strategy.differentiation:
            lines.append(f"- {item}")
        lines.append("")

    if strategy.target_focus:
        lines += ["## Focus sul target", "", strategy.target_focus, ""]

    if strategy.key_messages:
        lines += ["## Messaggi chiave", ""]
        for msg in strategy.key_messages:
            lines.append(f"- {msg}")
        lines.append("")

    if strategy.content_pillars:
        lines += ["## Pilastri di contenuto", ""]
        for pillar in strategy.content_pillars:
            lines.append(f"### {pillar.title}")
            if pillar.description:
                lines.append(pillar.description)
            if pillar.sample_topics:
                lines.append("")
                for topic in pillar.sample_topics:
                    lines.append(f"- {topic}")
            lines.append("")

    if strategy.channel_strategy:
        lines += ["## Strategia di canale", ""]
        for item in strategy.channel_strategy:
            lines.append(f"- {item}")
        lines.append("")

    if strategy.recommended_actions:
        lines += ["## Azioni raccomandate", ""]
        for i, action in enumerate(strategy.recommended_actions, 1):
            lines.append(f"{i}. {action}")
        lines.append("")

    if strategy.kpis:
        lines += ["## KPI e metriche di successo", ""]
        for kpi in strategy.kpis:
            lines.append(f"- {kpi}")
        lines.append("")

    if strategy.editorial_plan:
        lines += ["## Piano editoriale", ""]
        current_week = 0
        for item in strategy.editorial_plan:
            if item.week != current_week:
                current_week = item.week
                lines += [f"### Settimana {current_week}", ""]
                lines += ["| Giorno | Piattaforma | Formato | Pilastro | Idea / Hook | Obiettivo |"]
                lines += ["|--------|-------------|---------|----------|-------------|-----------|"]
            lines.append(
                f"| {item.day} | {item.platform} | {item.format} "
                f"| {item.pillar} | {item.topic} | {item.goal} |"
            )
        lines.append("")

    return "\n".join(lines)


def export_excel(rows: list[CompetitorRow], output_path: Path) -> None:
    """Export competitor rows to an Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    if not rows:
        return

    flat_rows = [_flatten(r) for r in rows]
    headers = list(flat_rows[0].keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Competitor Analysis"

    # Header row with Kolif Agency brand colors
    header_fill = PatternFill(start_color="09084C", end_color="09084C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    accent_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    for row_idx, row_data in enumerate(flat_rows, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = accent_fill

    # Auto-size columns (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 40)

    ws.row_dimensions[1].height = 30
    wb.save(output_path)
